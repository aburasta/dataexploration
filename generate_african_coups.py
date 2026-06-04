#!/usr/bin/env python3
"""
Generate spreadsheets of African coups d'etat (1950-present).

Data is compiled from the historical record following the classification
conventions of the Powell & Thyne Coup d'Etat Dataset:
  - Successful  : perpetrators seized and held power for at least ~7 days
  - Unsuccessful: failed / attempted coups (including thwarted plots and
                  short-lived takeovers that were quickly reversed)

NOTE ON PROVENANCE: The authoritative Powell & Thyne data file could not be
downloaded automatically from this sandbox (network allowlist). The records
below are a best-effort compilation of well-documented African coup events.
To use the official dataset, allowlist www.uky.edu / jonathanmpowell.com and
re-run, or drop the official .txt in this folder and we can parse it directly.

Outputs (both import cleanly into Google Sheets via File -> Import -> Upload):
  - african_coups_1950_present.xlsx   (tabs: All, Successful, Unsuccessful)
  - african_coups_unsuccessful.xlsx   (the unsuccessful list only)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# (country, date, outcome, target/leader, note)
# date is "YYYY-MM-DD" where confidently known, otherwise "YYYY".
# outcome: "Successful" or "Unsuccessful"
COUPS = [
    # ---------------- NORTH AFRICA ----------------
    ("Egypt", "1952-07-23", "Successful", "King Farouk", "Free Officers (Naguib/Nasser)"),
    ("Egypt", "2013-07-03", "Successful", "Mohamed Morsi", "el-Sisi removes elected president"),
    ("Sudan", "1958-11-17", "Successful", "Abdullah Khalil govt", "Ibrahim Abboud"),
    ("Sudan", "1969-05-25", "Successful", "Ismail al-Azhari", "Gaafar Nimeiry"),
    ("Sudan", "1971-07-19", "Unsuccessful", "Nimeiry", "Hashem al-Atta; reversed 3 days later"),
    ("Sudan", "1976-07-02", "Unsuccessful", "Nimeiry", "Sadiq al-Mahdi-linked attempt"),
    ("Sudan", "1985-04-06", "Successful", "Nimeiry", "Abdel Rahman Swar al-Dahab"),
    ("Sudan", "1989-06-30", "Successful", "Sadiq al-Mahdi", "Omar al-Bashir"),
    ("Sudan", "1990-04-23", "Unsuccessful", "al-Bashir", "officers executed"),
    ("Sudan", "2019-04-11", "Successful", "Omar al-Bashir", "military removes Bashir after uprising"),
    ("Sudan", "2021-09-21", "Unsuccessful", "transitional govt", "Bashir-loyalist attempt"),
    ("Sudan", "2021-10-25", "Successful", "transitional govt", "Abdel Fattah al-Burhan"),
    ("Libya", "1969-09-01", "Successful", "King Idris I", "Muammar Gaddafi"),
    ("Libya", "1975-08", "Unsuccessful", "Gaddafi", "officers' plot"),
    ("Libya", "1993-10", "Unsuccessful", "Gaddafi", "Warfalla-linked attempt"),
    ("Algeria", "1965-06-19", "Successful", "Ahmed Ben Bella", "Houari Boumediene"),
    ("Tunisia", "1987-11-07", "Successful", "Habib Bourguiba", "Ben Ali ('medical coup')"),
    ("Morocco", "1971-07-10", "Unsuccessful", "King Hassan II", "Skhirat palace attack"),
    ("Morocco", "1972-08-16", "Unsuccessful", "King Hassan II", "air force attack (Oufkir)"),
    ("Mauritania", "1978-07-10", "Successful", "Moktar Ould Daddah", "CMRN"),
    ("Mauritania", "1979-04-06", "Successful", "Mustafa Ould Salek", "Ould Bouceif"),
    ("Mauritania", "1980-01-04", "Successful", "Ould Louly", "Haidalla"),
    ("Mauritania", "1981-03-16", "Unsuccessful", "Haidalla", "attempted coup"),
    ("Mauritania", "1984-12-12", "Successful", "Haidalla", "Maaouya Ould Taya"),
    ("Mauritania", "2003-06-08", "Unsuccessful", "Taya", "Knights of Change attempt"),
    ("Mauritania", "2005-08-03", "Successful", "Taya", "Ely Ould Mohamed Vall"),
    ("Mauritania", "2008-08-06", "Successful", "Abdallahi", "Abdel Aziz"),

    # ---------------- WEST AFRICA ----------------
    ("Nigeria", "1966-01-15", "Successful", "Tafawa Balewa govt", "Nzeogwu; led to Ironsi"),
    ("Nigeria", "1966-07-29", "Successful", "Aguiyi-Ironsi", "counter-coup; Gowon"),
    ("Nigeria", "1975-07-29", "Successful", "Yakubu Gowon", "Murtala Mohammed"),
    ("Nigeria", "1976-02-13", "Unsuccessful", "Murtala/Obasanjo", "Dimka; Murtala killed"),
    ("Nigeria", "1983-12-31", "Successful", "Shehu Shagari", "Muhammadu Buhari"),
    ("Nigeria", "1985-08-27", "Successful", "Buhari", "Ibrahim Babangida"),
    ("Nigeria", "1990-04-22", "Unsuccessful", "Babangida", "Gideon Orkar"),
    ("Nigeria", "1993-11-17", "Successful", "Shonekan interim govt", "Sani Abacha"),
    ("Ghana", "1966-02-24", "Successful", "Kwame Nkrumah", "NLC"),
    ("Ghana", "1967-04-17", "Unsuccessful", "NLC", "Lt. Arthur"),
    ("Ghana", "1972-01-13", "Successful", "Kofi Busia", "Acheampong"),
    ("Ghana", "1978-07-05", "Successful", "Acheampong", "palace coup; Akuffo"),
    ("Ghana", "1979-06-04", "Successful", "Akuffo", "Jerry Rawlings"),
    ("Ghana", "1981-12-31", "Successful", "Hilla Limann", "Rawlings (2nd)"),
    ("Togo", "1963-01-13", "Successful", "Sylvanus Olympio", "assassinated; Grunitzky installed"),
    ("Togo", "1967-01-13", "Successful", "Grunitzky", "Gnassingbe Eyadema"),
    ("Benin", "1963-10-28", "Successful", "Hubert Maga", "Christophe Soglo"),
    ("Benin", "1965-11-29", "Successful", "Sourou Apithy/Ahomadegbe", "Soglo"),
    ("Benin", "1967-12-17", "Successful", "Soglo", "Maurice Kouandete"),
    ("Benin", "1969-12-10", "Successful", "Zinsou", "military"),
    ("Benin", "1972-10-26", "Successful", "Presidential Council", "Mathieu Kerekou"),
    ("Benin", "1977-01-16", "Unsuccessful", "Kerekou", "Bob Denard mercenary raid"),
    ("Benin", "1988", "Unsuccessful", "Kerekou", "attempted coup"),
    ("Burkina Faso", "1966-01-03", "Successful", "Maurice Yameogo", "Sangoule Lamizana"),
    ("Burkina Faso", "1980-11-25", "Successful", "Lamizana", "Saye Zerbo"),
    ("Burkina Faso", "1982-11-07", "Successful", "Zerbo", "Jean-Baptiste Ouedraogo"),
    ("Burkina Faso", "1983-08-04", "Successful", "Ouedraogo", "Thomas Sankara"),
    ("Burkina Faso", "1987-10-15", "Successful", "Sankara", "Blaise Compaore; Sankara killed"),
    ("Burkina Faso", "2014-11-01", "Successful", "Compaore", "military takeover after uprising"),
    ("Burkina Faso", "2015-09-16", "Unsuccessful", "transitional govt", "RSP (Diendere); reversed"),
    ("Burkina Faso", "2022-01-24", "Successful", "Roch Kabore", "Paul-Henri Damiba"),
    ("Burkina Faso", "2022-09-30", "Successful", "Damiba", "Ibrahim Traore"),
    ("Mali", "1968-11-19", "Successful", "Modibo Keita", "Moussa Traore"),
    ("Mali", "1991-03-26", "Successful", "Moussa Traore", "Amadou Toumani Toure"),
    ("Mali", "2012-03-22", "Successful", "Amadou Toumani Toure", "Amadou Sanogo"),
    ("Mali", "2012-04-30", "Unsuccessful", "junta", "counter-coup attempt"),
    ("Mali", "2020-08-18", "Successful", "Ibrahim Boubacar Keita", "CNSP / Goita"),
    ("Mali", "2021-05-24", "Successful", "Bah Ndaw transition", "Assimi Goita"),
    ("Niger", "1974-04-15", "Successful", "Hamani Diori", "Seyni Kountche"),
    ("Niger", "1996-01-27", "Successful", "Mahamane Ousmane", "Ibrahim Bare Mainassara"),
    ("Niger", "1999-04-09", "Successful", "Bare Mainassara", "assassinated; Wanke"),
    ("Niger", "2010-02-18", "Successful", "Mamadou Tandja", "Salou Djibo"),
    ("Niger", "2021-03-31", "Unsuccessful", "Bazoum (incoming)", "days before inauguration"),
    ("Niger", "2023-07-26", "Successful", "Mohamed Bazoum", "Abdourahamane Tchiani"),
    ("Guinea", "1984-04-03", "Successful", "Beavogui interim", "Lansana Conte (after Toure death)"),
    ("Guinea", "1985-07-04", "Unsuccessful", "Conte", "Diarra Traore"),
    ("Guinea", "2008-12-23", "Successful", "interim (Conte died)", "Moussa Dadis Camara"),
    ("Guinea", "2021-09-05", "Successful", "Alpha Conde", "Mamady Doumbouya"),
    ("Guinea-Bissau", "1980-11-14", "Successful", "Luis Cabral", "Joao Bernardo Vieira"),
    ("Guinea-Bissau", "1999-05-07", "Successful", "Vieira", "military / Mane"),
    ("Guinea-Bissau", "2003-09-14", "Successful", "Kumba Iala", "Verissimo Seabra"),
    ("Guinea-Bissau", "2012-04-12", "Successful", "interim (pre-runoff)", "military junta"),
    ("Guinea-Bissau", "2022-02-01", "Unsuccessful", "Umaro Sissoco Embalo", "attempted coup"),
    ("Guinea-Bissau", "2023-12-01", "Unsuccessful", "Embalo", "clashes / attempted coup"),
    ("Sierra Leone", "1967-03-23", "Successful", "Siaka Stevens (elect)", "Lansana / NRC"),
    ("Sierra Leone", "1968-04-18", "Successful", "NRC", "counter-coup restored Stevens"),
    ("Sierra Leone", "1992-04-29", "Successful", "Joseph Momoh", "Valentine Strasser"),
    ("Sierra Leone", "1996-01-16", "Successful", "Strasser", "Julius Maada Bio (palace)"),
    ("Sierra Leone", "1997-05-25", "Successful", "Tejan Kabbah", "Johnny Paul Koroma (AFRC)"),
    ("Sierra Leone", "2023-11-26", "Unsuccessful", "Julius Maada Bio", "attempted coup / armoury raid"),
    ("Liberia", "1980-04-12", "Successful", "William Tolbert", "Samuel Doe; Tolbert killed"),
    ("Liberia", "1985-11-12", "Unsuccessful", "Samuel Doe", "Thomas Quiwonkpa"),
    ("Gambia", "1981-07-30", "Unsuccessful", "Dawda Jawara", "Kukoi Samba Sanyang; Senegal reversed"),
    ("Gambia", "1994-07-22", "Successful", "Dawda Jawara", "Yahya Jammeh"),
    ("Gambia", "2014-12-30", "Unsuccessful", "Yahya Jammeh", "attempted coup"),
    ("Ivory Coast", "1999-12-24", "Successful", "Henri Konan Bedie", "Robert Guei"),

    # ---------------- CENTRAL AFRICA ----------------
    ("Central African Republic", "1966-01-01", "Successful", "David Dacko", "Jean-Bedel Bokassa"),
    ("Central African Republic", "1979-09-20", "Successful", "Bokassa", "French-backed; Dacko restored"),
    ("Central African Republic", "1981-09-01", "Successful", "David Dacko", "Andre Kolingba"),
    ("Central African Republic", "2003-03-15", "Successful", "Ange-Felix Patasse", "Francois Bozize"),
    ("Central African Republic", "2013-03-24", "Successful", "Francois Bozize", "Seleka / Djotodia"),
    ("Chad", "1975-04-13", "Successful", "Francois Tombalbaye", "killed; Malloum"),
    ("Chad", "1982-06-07", "Successful", "Goukouni Oueddei", "Hissene Habre"),
    ("Chad", "1990-12-01", "Successful", "Hissene Habre", "Idriss Deby"),
    ("Cameroon", "1984-04-06", "Unsuccessful", "Paul Biya", "Republican Guard attempt"),
    ("Gabon", "1964-02-18", "Unsuccessful", "Leon Mba", "reversed by French intervention"),
    ("Gabon", "2019-01-07", "Unsuccessful", "Ali Bongo", "mutiny attempt"),
    ("Gabon", "2023-08-30", "Successful", "Ali Bongo", "Brice Oligui Nguema"),
    ("Republic of the Congo", "1968-08-03", "Successful", "Alphonse Massamba-Debat", "Marien Ngouabi"),
    ("DR Congo (Zaire)", "1960-09-14", "Successful", "Lumumba govt", "Mobutu (first)"),
    ("DR Congo (Zaire)", "1965-11-24", "Successful", "Kasavubu", "Mobutu (second)"),
    ("Equatorial Guinea", "1979-08-03", "Successful", "Francisco Macias Nguema", "Teodoro Obiang"),
    ("Equatorial Guinea", "2004-03-07", "Unsuccessful", "Obiang", "'Wonga coup' mercenary plot"),
    ("Sao Tome and Principe", "1995-08-15", "Unsuccessful", "Trovoada", "brief coup, reversed"),
    ("Sao Tome and Principe", "2003-07-16", "Unsuccessful", "Fradique de Menezes", "brief coup, reversed"),

    # ---------------- EAST AFRICA / HORN ----------------
    ("Ethiopia", "1960-12-13", "Unsuccessful", "Haile Selassie", "Imperial Guard attempt"),
    ("Ethiopia", "1974-09-12", "Successful", "Haile Selassie", "Derg"),
    ("Ethiopia", "1989-05-16", "Unsuccessful", "Mengistu", "officers' attempt"),
    ("Somalia", "1969-10-21", "Successful", "Sharmarke govt", "Siad Barre"),
    ("Somalia", "1978-04-09", "Unsuccessful", "Siad Barre", "attempted coup"),
    ("Uganda", "1971-01-25", "Successful", "Milton Obote", "Idi Amin"),
    ("Uganda", "1985-07-27", "Successful", "Milton Obote (2nd)", "Tito Okello"),
    ("Kenya", "1982-08-01", "Unsuccessful", "Daniel arap Moi", "air force attempt"),
    ("Burundi", "1966-07-08", "Successful", "Mwambutsa IV", "Ntare V (monarchical)"),
    ("Burundi", "1966-11-28", "Successful", "Ntare V", "Michel Micombero; monarchy abolished"),
    ("Burundi", "1976-11-01", "Successful", "Micombero", "Jean-Baptiste Bagaza"),
    ("Burundi", "1987-09-03", "Successful", "Bagaza", "Pierre Buyoya"),
    ("Burundi", "1993-10-21", "Unsuccessful", "Melchior Ndadaye", "killed; coup failed to hold"),
    ("Burundi", "1996-07-25", "Successful", "Ntibantunganya", "Pierre Buyoya (2nd)"),
    ("Burundi", "2015-05-13", "Unsuccessful", "Pierre Nkurunziza", "Godefroid Niyombare"),
    ("Rwanda", "1973-07-05", "Successful", "Gregoire Kayibanda", "Juvenal Habyarimana"),
    ("Comoros", "1975-08-03", "Successful", "Ahmed Abdallah", "Said Mohamed Jaffar"),
    ("Comoros", "1978-05-13", "Successful", "Ali Soilih", "Bob Denard; Abdallah restored"),
    ("Comoros", "1995-09-28", "Unsuccessful", "Said Mohamed Djohar", "Denard; reversed by France"),
    ("Comoros", "1999-04-30", "Successful", "Tadjidine Ben Said Massounde", "Azali Assoumani"),
    ("Madagascar", "2009-03-17", "Successful", "Marc Ravalomanana", "military-backed; Rajoelina"),
    ("Madagascar", "2010-11-17", "Unsuccessful", "Rajoelina", "officers' attempt"),
    ("Seychelles", "1977-06-05", "Successful", "James Mancham", "France-Albert Rene"),
    ("Seychelles", "1981-11-25", "Unsuccessful", "Rene", "Mike Hoare mercenary raid"),

    # ---------------- SOUTHERN AFRICA ----------------
    ("Lesotho", "1986-01-20", "Successful", "Leabua Jonathan", "Justin Lekhanya"),
    ("Lesotho", "1991-04-30", "Successful", "Lekhanya", "Elias Ramaema"),
    ("Lesotho", "2014-08-30", "Unsuccessful", "Tom Thabane", "attempted coup"),
    ("Zambia", "1990-06-30", "Unsuccessful", "Kenneth Kaunda", "brief announced coup"),
    ("Zambia", "1997-10-28", "Unsuccessful", "Frederick Chiluba", "Captain Solo"),
    ("Zimbabwe", "2017-11-15", "Successful", "Robert Mugabe", "military; Mugabe resigned"),
]


def parse_year(date_str):
    return int(date_str[:4])


HEADERS = ["#", "Country", "Date", "Year", "Outcome", "Deposed / Target", "Notes"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUCCESS_FILL = PatternFill("solid", fgColor="E2EFDA")
FAIL_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build_sheet(ws, rows, tint=None):
    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for i, (country, date, outcome, target, note) in enumerate(rows, start=1):
        ws.append([i, country, date, parse_year(date), outcome, target, note])
        r = ws.max_row
        fill = tint or (SUCCESS_FILL if outcome == "Successful" else FAIL_FILL)
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if c == 5:
                cell.fill = fill
    widths = [5, 26, 13, 7, 14, 26, 46]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"


def chrono(rows):
    return sorted(rows, key=lambda x: x[1])


all_rows = chrono(COUPS)
successful = chrono([c for c in COUPS if c[2] == "Successful"])
unsuccessful = chrono([c for c in COUPS if c[2] == "Unsuccessful"])

# ---- Spreadsheet 1: classified workbook ----
wb1 = Workbook()
ws_all = wb1.active
ws_all.title = "All Coups"
build_sheet(ws_all, all_rows)
build_sheet(wb1.create_sheet("Successful"), successful)
build_sheet(wb1.create_sheet("Unsuccessful"), unsuccessful)
wb1.save("african_coups_1950_present.xlsx")

# ---- Spreadsheet 2: unsuccessful only ----
wb2 = Workbook()
ws2 = wb2.active
ws2.title = "Unsuccessful Coups"
build_sheet(ws2, unsuccessful)
wb2.save("african_coups_unsuccessful.xlsx")

print(f"Total coup events : {len(all_rows)}")
print(f"  Successful       : {len(successful)}")
print(f"  Unsuccessful     : {len(unsuccessful)}")
print("Wrote: african_coups_1950_present.xlsx (tabs: All / Successful / Unsuccessful)")
print("Wrote: african_coups_unsuccessful.xlsx")
