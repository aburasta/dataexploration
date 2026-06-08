#!/usr/bin/env python3
"""Build the final all-in-one PRODUCTION MASTER spreadsheet.
Reads the existing matches workbook and adds: a derived primary/secondary match,
a core lesson, a benefit-led hook angle (per concept), and production-tracking
columns. One row per episode."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

SRC = "/home/user/dataexploration/finance_animal_behavior_matches.xlsx"
OUT = "/home/user/dataexploration/finance_production_master.xlsx"

# (lesson, hook) per concept #1..#100, in the same order as the matches file
HL = [
("Watch the crowd's signal, not the price — danger is when buying flips to selling.", "What if you could feel a crash turning a heartbeat before everyone else?"),
("Check the facts yourself; don't just copy the last person who acted.", "Why do smart crowds confidently walk off a cliff together?"),
("When the crowd's size is the only reason to buy, get cautious.", "How do bubbles trick careful people into buying at the very top?"),
("Exit before the door shrinks; crowded trades have narrow exits.", "Why can everyone sell today but no one sell tomorrow?"),
("Never short without a cap — being forced to cover fuels the fire.", "How can betting against something destroy you even when you're right?"),
("Extremes rarely last; fade the over-stretched, don't chase it.", "What if the safest profit is just waiting for things to snap back to normal?"),
("Trade less and use limit orders — the spread is a hidden toll.", "Who quietly takes a cut every time you buy or sell?"),
("Someone always bears the risk of holding inventory; know who and why.", "Why does the person selling you the stock want rid of it so fast?"),
("Big visible orders get picked off — size and timing matter.", "How do some traders profit from your order before it even lands?"),
("Your own size moves the price; scale in patiently.", "Why does buying a lot of something make it cost more as you go?"),
("Price is aggregated opinion, not truth — respect it but question it.", "How does a chaotic crowd magically agree on one fair price?"),
("In panic, safety gets crowded and expensive — own some beforehand.", "Where does all the money run the instant fear hits?"),
("If value depends on a greater fool, you might be it.", "What if the thing you bought is only worth what the next sucker pays?"),
("The bottom is mass exhaustion — opportunity hides in capitulation.", "How do you spot the exact moment a crash hits rock bottom?"),
("Not every bounce is a recovery — beware the dead-cat bounce.", "Why does a falling market jump right before it falls further?"),
("Pool risks too big to carry alone — that's insurance.", "What if one disaster could never wipe you out?"),
("Stock more than the average day needs; rare disasters are the killers.", "Why do the calmest people prepare for things that almost never happen?"),
("Accept a small certain cost to cap a catastrophic one.", "How do you pay a little now to never get destroyed later?"),
("Uncorrelated bets cut risk without cutting return — don't concentrate.", "Why is spreading your bets the only free lunch in finance?"),
("In a crisis everything moves together — plan for it.", "Why does diversification fail exactly when you need it most?"),
("Watch for incentives to take risks others will absorb.", "What happens when someone else pays for your mistakes?"),
("When only bad risks want the deal, the deal is mispriced.", "Why are the eagerest customers often the riskiest?"),
("Even your backstops need backstops — layer your protection.", "Who insures the insurance companies?"),
("Bearing a little yourself keeps the pool honest and cheap.", "Why does sharing a small loss protect everyone?"),
("A hedge that doesn't track your risk leaves a dangerous gap.", "What if your safety net doesn't quite match what it's protecting?"),
("Count the failures you can't see before copying winners.", "Why do success stories quietly lie to you?"),
("Avoid bets that can end the game — survival comes first.", "What's the one loss you can never come back from?"),
("Size bets to your real edge so bad luck can't ruin you.", "How much should you bet when you finally have an edge?"),
("Pooled money in the gap before claims is a quiet asset.", "How do insurers make money just holding your premiums?"),
("Keep an emergency reserve sized for the worst season.", "Why do the people who never panic always have a buffer?"),
("Compounding is slow then sudden — start early, stay in.", "When exactly does saving turn into real wealth?"),
("Interrupting compounding resets the engine — don't break the chain.", "Why does cashing out early cost far more than it looks?"),
("Reinvest the yield to grow a compounding asset.", "What if you fed your returns back instead of spending them?"),
("Use the Rule of 72 to see a small rate's real power.", "How fast does your money actually double?"),
("Losses compound too — protecting the downside matters most.", "Why does a 50% loss need a 100% gain to recover?"),
("Idle cash loses value — beat inflation or lose ground.", "Why is your cash quietly shrinking while you sleep?"),
("When trust breaks, value can spiral away fast.", "What makes money collapse into worthlessness overnight?"),
("Only growth after inflation is real growth.", "Are you actually richer, or just holding bigger numbers?"),
("Money now beats money later — value it accordingly.", "Why is a dollar today worth more than a dollar tomorrow?"),
("Distant rewards are worth less now — discount them.", "How do you put a price on a payoff years away?"),
("Leverage cuts both ways equally; respect the downside.", "How can borrowing multiply your wins — and your wipeouts?"),
("Escape compounding debt early before it traps you.", "Why do some debts grow faster than you can ever pay?"),
("Collateral lowers the lender's risk and your rate.", "Why does pledging what you own get you a better deal?"),
("Reliable repayment compounds into bigger future access.", "How do you turn a reputation into borrowing power?"),
("Know who you're connected to — failure spreads.", "How does one failure take down everyone nearby?"),
("Leverage can liquidate you involuntarily — keep a buffer.", "What forces you to sell at the worst possible moment?"),
("Know your place in the line before you lend or invest.", "When the money runs out, who actually gets paid first?"),
("Refinance when terms improve — don't overpay out of inertia.", "How do you swap a bad deal for a better one?"),
("Don't fund long commitments with money that can vanish.", "Why do healthy institutions suddenly collapse overnight?"),
("Pre-agreed rules keep both sides honest.", "How do lenders protect themselves before trouble even starts?"),
("Steady principal payments shrink the burden predictably.", "How does a huge debt quietly disappear over time?"),
("Never borrow from a position of desperation.", "Why do the people who can least afford it pay the most?"),
("Confidence is the asset — panics are self-fulfilling.", "How does the fear of a collapse cause the collapse?"),
("A credible safety net prevents the panic itself.", "Why does a backstop work best when it's never actually used?"),
("Banks multiply money — understand where it comes from.", "How does one deposit quietly become many loans?"),
("Circulation, not just quantity, drives the economy.", "Why does the same dollar make us richer the faster it moves?"),
("Watch the hubs — connected failures cascade.", "How does one domino topple the whole system?"),
("Removing the reason to panic prevents the panic.", "Why does a guarantee stop a bank run before it starts?"),
("Cheap money props up activity — and inflates assets.", "What really happens when central banks flood the system?"),
("One central lever moves every connected price.", "How does one rate change ripple all the way into your wallet?"),
("When fear dominates, easy money gets hoarded, not spent.", "Why does cheap money sometimes do nothing at all?"),
("A dominant borrower can starve everyone else of credit.", "How does big borrowing quietly squeeze out the little guy?"),
("The power to issue money is itself a source of wealth.", "Who profits simply from printing the money?"),
("Defending a fixed price drains finite reserves — it can break.", "Why do countries burn billions defending a single number?"),
("The carry works until it snaps — manage the reversal.", "How do you earn by borrowing cheap and lending dear?"),
("Deflation freezes spending and feeds on itself.", "Why can falling prices actually destroy an economy?"),
("Every yes is a no to something else — count it.", "What's the hidden price of every choice you make?"),
("Ignore what's spent; decide on the future only.", "Why do we keep pouring money into clearly losing bets?"),
("Scale lowers unit cost — size is an advantage.", "Why does getting bigger make everything cheaper?"),
("Value that grows with users builds a moat.", "Why do the biggest platforms become unstoppable?"),
("Durable advantages protect long-term returns.", "What actually keeps competitors from stealing your profits?"),
("Incumbents fall to nimble newcomers — adapt or fade.", "Why must the old be destroyed for the new to thrive?"),
("Balance paying out against reinvesting to grow.", "Harvest now, or grow a bigger harvest later?"),
("Fewer shares concentrate value in those that remain.", "How does a company make each share worth more by shrinking?"),
("Watch share count — issuance dilutes existing owners.", "How does new stock quietly shrink your slice?"),
("Value to current market price, not your purchase price.", "What is your asset really worth right now?"),
("Intangibles can carry real value — and real risk.", "Why pay far more than something is physically worth?"),
("A working whole beats the sum of sold-off pieces.", "Why is a living business worth more than its parts?"),
("Expensive, hard-to-fake signals are the honest ones.", "How do you prove a quality that can't be faked?"),
("Align incentives or monitor — interests drift.", "Why don't the people you hire always act in your interest?"),
("Don't let fear of loss drive bad decisions.", "Why does losing $100 hurt more than gaining $100 feels good?"),
("Invest steadily regardless of conditions.", "How do you invest without ever needing to time the market?"),
("Trade access for yield deliberately, not by accident.", "Why does locking your money up pay you more?"),
("Don't anchor on irrelevant reference prices.", "Why does the first number you see hijack your judgment?"),
("Popularity isn't proof — verify before you follow.", "Why do we instinctively trust what everyone else is buying?"),
("Don't extrapolate the latest trend forever.", "Why do we assume the recent past will just keep going?"),
("Cut losers, let winners run — fight the instinct.", "Why do we sell our winners and cling to our losers?"),
("Fear of missing out is the most expensive emotion.", "Why do we always seem to buy in right at the top?"),
("Patience compounds — delay gratification.", "Smaller reward now, or a bigger one if you can wait?"),
("Money is fungible — don't fool yourself with mental jars.", "Why do we treat the exact same dollar completely differently?"),
("Price gaps are free money — until they close.", "How do you profit from the same thing priced two ways?"),
("Seek capped downside with open-ended upside.", "How do you risk a little to win a lot?"),
("Options bleed with time — time itself is a cost.", "Why do some bets lose value just by waiting?"),
("Forced buying can self-accelerate violently.", "How does quiet hedging force a price to suddenly rocket?"),
("Turbulence clusters — expect storms after storms.", "Why do calm and chaos seem to arrive in streaks?"),
("The spread to spot signals scarcity or glut.", "What does the futures price secretly reveal about the future?"),
("A deal is only as good as who's on the other side.", "What if the other side of your deal simply can't pay?"),
("Offset mutual obligations to cut exposure.", "How do you settle many debts by moving almost nothing?"),
("Know which tranche you hold — order of losses matters.", "How do you slice one risk into safe and risky layers?"),
("Rebalance to your targets — sell high, buy low automatically.", "How does trimming your winners actually keep you disciplined?"),
]
assert len(HL) == 100

# ---- read source matches ----
swb = openpyxl.load_workbook(SRC)
sws = swb["Concept-Animal Matches"]
src = []
for r in range(2, 102):
    src.append([sws.cell(row=r, column=c).value for c in range(1, 12)])
assert len(src) == 100

# ---- build output ----
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Production Master"

headers = [
    "#", "Category", "Granular Concept", "The Specific Finance Aspect",
    "Core Lesson / Takeaway", "Suggested Hook Angle (benefit-led)",
    "Primary Animal Match", "Primary %", "Primary Behavior Description (script-ready)",
    "Secondary Animal Match", "Secondary %", "Secondary Behavior Description",
    "Notes / Verification",
    "Episode #", "Status", "Runtime", "Publish link / date",
]

hfill = PatternFill("solid", fgColor="1F4E5F")
hfont = Font(bold=True, color="FFFFFF", size=10)
catfill = PatternFill("solid", fgColor="EAF1F3")
prfill = PatternFill("solid", fgColor="F3F8E8")
scfill = PatternFill("solid", fgColor="FBF1E8")
trackfill = PatternFill("solid", fgColor="F0EEF6")
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(vertical="top", wrap_text=True)
ctr = Alignment(vertical="top", horizontal="center", wrap_text=True)


def qfill(q):
    try:
        q = int(q)
    except (TypeError, ValueError):
        return None
    if q >= 85: return PatternFill("solid", fgColor="C6EFCE")
    if q >= 75: return PatternFill("solid", fgColor="FFEB9C")
    return PatternFill("solid", fgColor="FFC7CE")


ws.append(headers)
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = hfill; cell.font = hfont
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    cell.border = border

for i, row in enumerate(src):
    n, cat, concept, aspect, m1, d1, q1, m2, d2, q2, notes = row
    lesson, hook = HL[i]
    # derive primary/secondary by quality
    if (q2 is not None and q1 is not None and q2 > q1):
        pm, pq, pd, scm, scq, scd = m2, q2, d2, m1, q1, d1
    else:
        pm, pq, pd, scm, scq, scd = m1, q1, d1, m2, q2, d2
    ws.append([n, cat, concept, aspect, lesson, hook,
               pm, pq, pd, scm, scq, scd, notes,
               "", "", "", ""])
    rr = i + 2
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=rr, column=c)
        cell.border = border
        cell.alignment = ctr if c in (1, 8, 11) else wrap
    ws.cell(row=rr, column=2).fill = catfill
    ws.cell(row=rr, column=3).font = Font(bold=True)
    ws.cell(row=rr, column=7).fill = prfill
    ws.cell(row=rr, column=7).font = Font(bold=True)
    ws.cell(row=rr, column=9).fill = prfill
    f = qfill(pq);  ws.cell(row=rr, column=8).fill = f if f else prfill
    ws.cell(row=rr, column=8).font = Font(bold=True)
    ws.cell(row=rr, column=10).fill = scfill
    ws.cell(row=rr, column=10).font = Font(bold=True)
    ws.cell(row=rr, column=12).fill = scfill
    f = qfill(scq); ws.cell(row=rr, column=11).fill = f if f else scfill
    for c in (14, 15, 16, 17):
        ws.cell(row=rr, column=c).fill = trackfill

# Status dropdown
dv = DataValidation(type="list",
                    formula1='"Planned,Scripted,Prompts ready,Images done,Edited,Published"',
                    allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"O2:O101")

widths = {"A": 4, "B": 20, "C": 26, "D": 34, "E": 34, "F": 36,
          "G": 22, "H": 8, "I": 50, "J": 22, "K": 9, "L": 50, "M": 22,
          "N": 9, "O": 14, "P": 9, "Q": 18}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = "C2"
ws.auto_filter.ref = f"A1:Q101"

# ---- How-to sheet ----
ws2 = wb.create_sheet("How to use")
rows2 = [
    ["FINANCE × ANIMAL BEHAVIOR — PRODUCTION MASTER", ""],
    ["", ""],
    ["This is the single input file for the series. One row = one episode.", ""],
    ["Pair it with PROJECT_MASTER_GUIDE.docx in your Claude Project.", ""],
    ["", ""],
    ["To make an episode:", "Tell Claude: 'Make the episode for row N' (or paste the row)."],
    ["You get back:", "A shot list (script + timestamps) + a Higgsfield prompt pack."],
    ["Then:", "Generate the images free in the Higgsfield web app (Seedream 4.5), edit, publish."],
    ["", ""],
    ["Columns", ""],
    ["E Core Lesson", "The practical takeaway the episode delivers."],
    ["F Hook Angle", "A benefit-led, suspenseful question to open the video (seed for the script)."],
    ["G-I Primary match", "The strongest animal pairing + its script-ready behavior description. Build the episode on this."],
    ["J-L Secondary match", "A backup/secondary facet, or use in the 'where it breaks down' beat."],
    ["H / K Quality %", "How cleanly the animal maps. Green >=85 strong, amber 75-84 ok, red <75 a stretch."],
    ["M Notes", "'Web-verified' = the biology was checked against scientific sources."],
    ["N-Q Tracking", "Fill these in as you produce: Episode #, Status (dropdown), Runtime, Publish link/date."],
    ["", ""],
    ["Suggested launch order (highest-confidence pairings):", ""],
    ["1) Risk pooling", "Vampire bat blood sharing (96%) — demo already built (EP07)."],
    ["2) Price discovery", "Honeybee swarm quorum (94%)."],
    ["3) Costly signaling", "Stotting gazelle / peacock (92%)."],
    ["4) Momentum / bubbles", "Desert locust phase change (92%)."],
    ["5) Creditworthiness", "Cleaner-fish image scoring (90%)."],
    ["6) Diversification", "Scatter-hoarding caches (90%)."],
]
for r in rows2:
    ws2.append(r)
ws2.column_dimensions["A"].width = 40
ws2.column_dimensions["B"].width = 75
ws2["A1"].font = Font(bold=True, size=13, color="1F4E5F")
for r in range(1, len(rows2) + 1):
    ws2.cell(row=r, column=2).alignment = wrap

wb.save(OUT)
print("Saved:", OUT)
print("Rows:", len(src), "| columns:", len(headers))
