#!/usr/bin/env python3
"""Build an Excel workbook of 100 granular finance concepts for the
animal-behavior animation series."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# (Category, Granular Concept, The specific aspect that maps to a behavior)
DATA = [
    # ---- Markets, microstructure & crashes ----
    ("Markets & Microstructure", "The inversion instant in a crash", "The exact moment the crowd's 'buy because others buy' signal flips to 'sell because others sell.'"),
    ("Markets & Microstructure", "Information cascades", "The point where people stop using their own private information and simply copy the last actor's choice."),
    ("Markets & Microstructure", "Momentum as self-fuel", "Rising prices recruiting new buyers, where the crowd's growing size is itself the reason to join."),
    ("Markets & Microstructure", "Liquidity evaporation", "Everyone reaching for the same exit at once physically shrinking the exit."),
    ("Markets & Microstructure", "The short squeeze", "Trapped sellers being forced to buy back, pouring fuel on the very fire burning them."),
    ("Markets & Microstructure", "Mean reversion", "An over-stretched price snapping back toward its long-run baseline."),
    ("Markets & Microstructure", "The bid-ask spread", "A tiny toll charged for providing immediate availability to whoever wants in or out now."),
    ("Markets & Microstructure", "Market-maker inventory risk", "Being stuck holding stock you must offload before its value drifts against you."),
    ("Markets & Microstructure", "Front-running", "Acting an instant ahead of a large known incoming order to skim the move."),
    ("Markets & Microstructure", "Slippage on large orders", "Your own buying pushing the price up against you as the order fills."),
    ("Markets & Microstructure", "Price discovery", "Thousands of tiny independent signals aggregating into one shared number."),
    ("Markets & Microstructure", "Flight to quality", "Capital abruptly fleeing risky assets and crowding into the few safest ones during panic."),
    ("Markets & Microstructure", "The greater-fool dynamic", "Buying something overpriced solely because you expect a more eager buyer behind you."),
    ("Markets & Microstructure", "Capitulation", "The final exhausted mass sell-off that empties out the last holders and marks the bottom."),
    ("Markets & Microstructure", "The dead-cat bounce", "A brief deceptive recovery inside a larger ongoing decline."),

    # ---- Risk & insurance ----
    ("Risk & Insurance", "Risk pooling", "Individual catastrophic loss absorbed by the collective surplus of the group."),
    ("Risk & Insurance", "Tail-risk preparation", "Stockpiling against the rare, high-cost event you almost never actually witness."),
    ("Risk & Insurance", "Hedging", "Accepting a small certain cost now to cap a large uncertain loss later."),
    ("Risk & Insurance", "The diversification effect", "Spreading across uncorrelated bets to cut volatility without cutting expected return."),
    ("Risk & Insurance", "Correlation breakdown in a crisis", "Things that normally move independently suddenly all moving together at the worst moment."),
    ("Risk & Insurance", "Moral hazard", "Taking on more risk precisely because someone else will absorb the downside."),
    ("Risk & Insurance", "Adverse selection", "The riskiest participants being the most eager to join the pool."),
    ("Risk & Insurance", "Reinsurance", "Insurers insuring each other so no single one can be wiped out by one event."),
    ("Risk & Insurance", "Deductibles", "Making the insured share the small losses so they don't make frivolous claims."),
    ("Risk & Insurance", "Basis risk", "A hedge that almost-but-not-quite tracks the thing it was meant to protect."),
    ("Risk & Insurance", "Survivorship bias", "Judging strategy by only the winners that survived, ignoring the silent dead."),
    ("Risk & Insurance", "Ruin risk", "A loss so total that you can never get back to the table to play again."),
    ("Risk & Insurance", "Kelly bet sizing", "Scaling each bet to your real edge so a streak of bad luck can't ruin you."),
    ("Risk & Insurance", "Insurance float", "Holding and using others' premiums during the gap before claims come due."),
    ("Risk & Insurance", "Precautionary reserves", "Holding idle surplus you hope never to use, sized to the worst plausible season."),

    # ---- Compounding, growth & erosion ----
    ("Compounding & Growth", "The compounding acceleration phase", "The late stage where growth visibly feeds on prior growth and the curve bends upward."),
    ("Compounding & Growth", "Breaking the compounding chain", "How one early withdrawal resets the whole engine rather than just denting it."),
    ("Compounding & Growth", "Reinvestment of yield", "Redirecting the harvest back into the principal instead of consuming it."),
    ("Compounding & Growth", "Doubling time (Rule of 72)", "A small steady rate quietly converting into a predictable doubling interval."),
    ("Compounding & Growth", "Negative compounding (drawdowns)", "Losses compounding downward, so recovery needs a bigger gain than the fall."),
    ("Compounding & Growth", "Inflation erosion", "The slow, continuous, barely noticed shrinkage of stored value."),
    ("Compounding & Growth", "Hyperinflation", "Value collapse feeding on itself as everyone spends faster to outrun it."),
    ("Compounding & Growth", "Real vs nominal return", "The true growth left only after the silent erosion is subtracted."),
    ("Compounding & Growth", "Time value of money", "A unit held now being worth more than the same unit promised later."),
    ("Compounding & Growth", "Discounting future cash flows", "Distant rewards shrinking in present worth the further away they sit."),

    # ---- Debt, credit & leverage ----
    ("Debt, Credit & Leverage", "Leverage's symmetric amplification", "Borrowed resources magnifying the gain and the fall by the exact same factor."),
    ("Debt, Credit & Leverage", "The debt spiral", "Interest compounding against you faster than you can possibly repay it."),
    ("Debt, Credit & Leverage", "Collateral", "Pledging something you value to shrink the lender's risk and your rate."),
    ("Debt, Credit & Leverage", "Creditworthiness as reputation", "Repeated reliable repayment slowly unlocking larger future access."),
    ("Debt, Credit & Leverage", "Default contagion", "One borrower's failure cascading through everyone financially connected to them."),
    ("Debt, Credit & Leverage", "The margin call", "Being forced to liquidate the instant your posted collateral drops too far."),
    ("Debt, Credit & Leverage", "Seniority in bankruptcy", "A strict pecking order deciding who eats first when the remaining assets are scarce."),
    ("Debt, Credit & Leverage", "Refinancing", "Swapping old expensive debt for cheaper new debt when conditions improve."),
    ("Debt, Credit & Leverage", "Maturity mismatch", "Borrowing short-term to fund long-term commitments, fine until everyone wants out."),
    ("Debt, Credit & Leverage", "Covenants", "Pre-agreed constraints the borrower accepts to reassure the lender."),
    ("Debt, Credit & Leverage", "Amortization", "Steadily chipping the principal down so the burden shrinks over time."),
    ("Debt, Credit & Leverage", "Desperation pricing (usury)", "Charging the trapped and desperate the most, because they cannot walk away."),

    # ---- Banking, macro & monetary ----
    ("Banking & Macro", "The bank run", "A self-fulfilling panic where the fear that others will withdraw causes the withdrawal."),
    ("Banking & Macro", "Lender of last resort", "A backstop that calms panic mostly by existing, often without ever being used."),
    ("Banking & Macro", "Fractional-reserve money creation", "The same deposited unit being lent out again and again, multiplying the supply."),
    ("Banking & Macro", "Money velocity", "The same money changing hands faster, multiplying its real effect on the economy."),
    ("Banking & Macro", "Systemic / network risk", "A local failure propagating through interconnected nodes instead of staying contained."),
    ("Banking & Macro", "Deposit insurance", "A guarantee that prevents the run precisely by making the run pointless."),
    ("Banking & Macro", "Quantitative easing", "Flooding the system with liquidity to force activity when normal channels stall."),
    ("Banking & Macro", "Interest-rate transmission", "One central rate change rippling outward through every connected price."),
    ("Banking & Macro", "The liquidity trap", "Easing stops working because fear keeps everyone hoarding no matter how cheap money gets."),
    ("Banking & Macro", "Crowding out", "A large dominant borrower (the state) displacing smaller private borrowers from the pool."),
    ("Banking & Macro", "Seigniorage", "The profit captured simply from being the one allowed to issue the money."),
    ("Banking & Macro", "Currency-peg defense", "Burning down a finite reserve to hold a price the market keeps pushing against."),
    ("Banking & Macro", "The carry trade", "Borrowing where it's cheap to deploy where it yields more, vulnerable to a sudden snap."),
    ("Banking & Macro", "The deflationary spiral", "Falling prices freezing spending, which lowers prices further in a self-reinforcing loop."),

    # ---- Corporate finance & valuation ----
    ("Corporate Finance & Valuation", "Opportunity cost", "The unseen cost of the path not taken whenever a finite resource is committed."),
    ("Corporate Finance & Valuation", "The sunk-cost fallacy", "Continuing because of what's already spent rather than what's still ahead."),
    ("Corporate Finance & Valuation", "Economies of scale", "Per-unit cost falling as size grows, rewarding the big over the small."),
    ("Corporate Finance & Valuation", "Network effects", "Each new participant making the whole more valuable to every existing one."),
    ("Corporate Finance & Valuation", "The economic moat", "Defending captured territory or resources to keep rivals from eroding your returns."),
    ("Corporate Finance & Valuation", "Creative destruction", "Nimble newcomers displacing entrenched incumbents that stopped adapting."),
    ("Corporate Finance & Valuation", "Dividend vs reinvestment", "Choosing between harvesting now and feeding the engine to grow the harvest later."),
    ("Corporate Finance & Valuation", "Share buybacks", "Shrinking the total pool so each remaining slice represents more."),
    ("Corporate Finance & Valuation", "Dilution", "New shares quietly shrinking the size of every existing holder's slice."),
    ("Corporate Finance & Valuation", "Mark-to-market", "Constantly revaluing what you hold to whatever it would fetch right now."),
    ("Corporate Finance & Valuation", "Goodwill", "Paying above tangible worth for something intangible you believe is really there."),
    ("Corporate Finance & Valuation", "Going-concern vs liquidation value", "A living, working whole being worth far more than the sum of its sold-off parts."),
    ("Corporate Finance & Valuation", "Costly signaling", "Spending real, hard-to-fake resources purely to prove a hidden quality is genuine."),
    ("Corporate Finance & Valuation", "The principal-agent problem", "A hired manager's incentives quietly drifting away from the owner's interests."),

    # ---- Behavioral finance & strategy ----
    ("Behavioral & Strategy", "Loss aversion", "Feeling an equivalent loss far more intensely than the matching gain."),
    ("Behavioral & Strategy", "Dollar-cost averaging", "Steady accumulation regardless of conditions to neutralize the timing problem."),
    ("Behavioral & Strategy", "Liquidity-vs-yield tradeoff", "Giving up instant access in exchange for a higher return on what's locked away."),
    ("Behavioral & Strategy", "Anchoring", "Fixating on an irrelevant first reference price and judging everything against it."),
    ("Behavioral & Strategy", "Social proof in adoption", "Trusting that if many others chose it, it must be safe to choose too."),
    ("Behavioral & Strategy", "Recency bias", "Over-weighting the most recent events as if they'll simply continue forever."),
    ("Behavioral & Strategy", "The disposition effect", "Selling winners too early while clinging to losers far too long."),
    ("Behavioral & Strategy", "FOMO-driven entry", "Buying late and high purely from the fear of being left behind."),
    ("Behavioral & Strategy", "Time preference", "Choosing a smaller reward now over a larger reward that requires waiting."),
    ("Behavioral & Strategy", "Mental accounting", "Treating identical money differently depending on which mental 'jar' it sits in."),

    # ---- Derivatives & advanced ----
    ("Derivatives & Advanced", "Arbitrage", "Exploiting the same thing being priced differently in two places at once."),
    ("Derivatives & Advanced", "Convexity / optionality", "A payoff with a capped downside but an open-ended upside."),
    ("Derivatives & Advanced", "Theta (time) decay", "An option steadily bleeding value with each day that passes, all else equal."),
    ("Derivatives & Advanced", "The gamma squeeze", "Hedging flows being forced to chase the underlying, accelerating its own move."),
    ("Derivatives & Advanced", "Volatility clustering", "Turbulent periods bunching together, calm following calm and storm following storm."),
    ("Derivatives & Advanced", "Contango / backwardation", "The future price sitting above or below today's, revealing scarcity or glut."),
    ("Derivatives & Advanced", "Counterparty risk", "The other side of your deal failing to pay when it finally matters."),
    ("Derivatives & Advanced", "Netting", "Offsetting mutual obligations so only the small net difference must actually change hands."),
    ("Derivatives & Advanced", "Tranching", "Slicing one pool of risk into stacked layers that absorb losses in strict order."),
    ("Derivatives & Advanced", "Portfolio rebalancing", "Periodically trimming what grew and topping up what shrank to restore target weights."),
]

assert len(DATA) == 100, f"Expected 100, got {len(DATA)}"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Finance Concepts"

headers = [
    "#", "Category", "Granular Concept",
    "The Specific Aspect That Maps",
    "Candidate Animal Behavior", "Mapping Strength (1-5)", "Notes",
]

# Styling
header_fill = PatternFill("solid", fgColor="1F4E5F")
header_font = Font(bold=True, color="FFFFFF", size=11)
cat_fill = PatternFill("solid", fgColor="EAF1F3")
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(vertical="top", wrap_text=True)
center = Alignment(vertical="top", horizontal="center")

ws.append(headers)
for col in range(1, len(headers) + 1):
    c = ws.cell(row=1, column=col)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    c.border = border

for i, (cat, concept, aspect) in enumerate(DATA, start=1):
    ws.append([i, cat, concept, aspect, "", "", ""])
    r = i + 1
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=r, column=col)
        cell.border = border
        cell.alignment = center if col in (1, 6) else wrap
    ws.cell(row=r, column=2).fill = cat_fill
    ws.cell(row=r, column=3).font = Font(bold=True)

# Column widths
widths = {"A": 5, "B": 26, "C": 32, "D": 60, "E": 30, "F": 16, "G": 28}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:G{len(DATA) + 1}"

out = "/home/user/dataexploration/finance_concepts_for_animation.xlsx"
wb.save(out)
print("Saved:", out)
print("Rows:", len(DATA))
# quick category tally
from collections import Counter
for cat, n in Counter(c for c, _, _ in DATA).items():
    print(f"  {n:>3}  {cat}")
